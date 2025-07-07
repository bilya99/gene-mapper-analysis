import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def gene_mapper_analysis(input_file_path, config_file_path, output_file_path, X=4, Y=8):
    """
    Analyze gene mapper data and create a formatted Excel output with conditional formatting.
    
    Args:
        input_file_path (str): Path to the input files directory
        config_file_path (str): Path to the config files directory  
        output_file_path (str): Path to the output files directory
        X (int): X value for EMX_MEY format (default: 4)
        Y (int): Y value for EMX_MEY format (default: 8)
    
    Returns:
        str: Path to the created output file
    """
    # User input for EMX_MEY format
    emx_mey_label = f"EM{X}_ME{Y}"

    # Load exported table and sample order
    df_exported_table = pd.read_excel(f"{input_file_path}/{emx_mey_label}_peaks_exported.xlsx")
    # After reading the DataFrame:
    df_exported_table.columns = df_exported_table.columns.str.strip()
    df_exported_table["Size"] = pd.to_numeric(df_exported_table["Size"], errors="coerce")
    # print(df_exported_table.columns)
    df_sample_order = pd.read_excel(f"{input_file_path}/Hyssop_sample_numbers_Genemapper.xlsx")

    # Create and save empty size ranges file if needed
    size_ranges_path = f"{input_file_path}/Size_ranges_{emx_mey_label}.xlsx"
    if not pd.io.common.file_exists(size_ranges_path):
        pd.DataFrame([["A", "B", "Lower", "Upper"]]).to_excel(size_ranges_path, header=False, index=False)

    df_exported_table['Sample File Name'] = df_exported_table['Sample File Name'].str.replace(' ', '', regex=False)   

    # Alternative approach using merge if you prefer:
    df_exported_table = df_exported_table.merge(
        df_sample_order[['Default_name', 'Sample_number']], 
        left_on='Sample File Name', 
        right_on='Default_name', 
        how='left'
    )
    df_exported_table['Sample File Name'] = df_exported_table['Sample_number'].fillna(df_exported_table['Sample File Name'])
    df_exported_table = df_exported_table.drop(['Default_name', 'Sample_number'], axis=1)

    # Save the modified dataframe to Excel
    df_exported_table.to_excel(f"{config_file_path}/Exported_table_changed_sample_names_{emx_mey_label}.xlsx", index=False)

    # Step 1: Load the size ranges from Excel file
    size_ranges = pd.read_excel(os.path.join(input_file_path, f"Size_ranges_{emx_mey_label}.xlsx"))

    # Step 2: Create the filtered dataframe first (all rows that match ANY criteria)
    # First filter by Height
    height_filtered = df_exported_table[df_exported_table['Height'] > 160]

    # Create a function to check if a size falls within any of the ranges
    def size_in_any_range(size_value):
        """Check if size falls within any of the defined ranges"""
        return ((size_value >= size_ranges['Lower']) & (size_value <= size_ranges['Upper'])).any()

    # Apply the size range filter
    # Using apply with the function to check each Size value
    size_mask = height_filtered['Size'].apply(size_in_any_range)
    df_filtered = height_filtered[size_mask].copy()

    # Step 3: Add the classification column
    df_filtered['Classification'] = np.nan  # Initialize with NaN values

    # Step 4: Loop through each range and assign classifications
    for i in range(len(size_ranges)):
        lower_bound = size_ranges.iloc[i]['Lower']
        upper_bound = size_ranges.iloc[i]['Upper']
        a_value = size_ranges.iloc[i]['A']
        b_value = size_ranges.iloc[i]['B']
        
        # Create condition for this range
        in_range = (df_filtered['Size'] >= lower_bound) & (df_filtered['Size'] <= upper_bound)
        
        # Assign classification based on A/B columns
        if a_value != 0 and b_value == 0:
            # Size is in column A, assign "c"
            df_filtered.loc[in_range, 'Classification'] = "c"
        elif b_value != 0 and a_value == 0:
            # Size is in column B, assign "b"
            df_filtered.loc[in_range, 'Classification'] = "b"

    # Create the new dataframe structure
    # First, get unique sample numbers to use as column names
    sample_numbers = df_sample_order['Sample_number'].unique().tolist()

    # Create a dataframe with EMX_MEY column and size ranges as rows
    df_matrix = pd.DataFrame({
        'EMX_MEY': [emx_mey_label] * len(size_ranges),  # Add EMX_MEY column
        'Size_Range': size_ranges['Lower'].astype(str) + " - " + size_ranges['Upper'].astype(str),
        'Lower': size_ranges['Lower'],
        'Upper': size_ranges['Upper']
    })

    # Add columns for each sample number, initialized with NaN
    for sample in sample_numbers:
        df_matrix[sample] = np.nan

    # Fill in the matrix with classification values
    for i in range(len(df_matrix)):
        lower_bound = df_matrix.iloc[i]['Lower']
        upper_bound = df_matrix.iloc[i]['Upper']
        
        # Find all filtered data that falls within this size range
        range_data = df_filtered[
            (df_filtered['Size'] >= lower_bound) & 
            (df_filtered['Size'] <= upper_bound)
        ]
        
        # For each sample in this size range, add the classification
        for j in range(len(range_data)):
            sample_name = range_data.iloc[j]['Sample File Name']
            classification = range_data.iloc[j]['Classification']
            
            # Check if this sample_name exists as a column in our matrix
            if sample_name in df_matrix.columns:
                df_matrix.loc[i, sample_name] = classification

    # Remove the helper columns (Lower, Upper) but keep EMX_MEY and Size_Range
    df_matrix_final = df_matrix.drop(['Lower', 'Upper'], axis=1)

    # Fill empty cells based on row content
    for i in range(len(df_matrix_final)):
        # Get all values in this row (excluding the EMX_MEY and Size_Range columns)
        row_values = df_matrix_final.iloc[i, 2:]  # Exclude first two columns
        
        # Check if row contains "c" or "b"
        has_c = (row_values == "c").any()
        has_b = (row_values == "b").any()
        
        # Fill NaN values based on what's present in the row
        if has_c:
            # If row contains "c", fill empty cells with "a"
            df_matrix_final.iloc[i, 2:] = df_matrix_final.iloc[i, 2:].fillna("a")
        elif has_b:
            # If row contains "b", fill empty cells with "d"
            df_matrix_final.iloc[i, 2:] = df_matrix_final.iloc[i, 2:].fillna("d")

    # View the result
    print("Matrix created:")
    print(df_matrix_final)

    # Export to Excel with conditional formatting
    print("\nCreating formatted Excel file...")

    # Create a formatted version with empty columns every 5 data columns
    df_formatted = df_matrix_final.copy()

    # Get column names (excluding the first two columns: EMX_MEY and Size_Range)
    data_columns = df_formatted.columns[2:].tolist()

    # Create new column structure with empty columns every 5
    new_columns = ["EMX_MEY", "Size_Range"]  # Start with EMX_MEY and Size_Range columns
    for i, col in enumerate(data_columns):
        new_columns.append(col)
        # Add empty column every 5 data columns (but not after the last column)
        if (i + 1) % 5 == 0 and i < len(data_columns) - 1:
            new_columns.append("")

    # Create the new formatted dataframe
    df_formatted_final = pd.DataFrame(index=df_formatted.index, columns=new_columns)

    # Copy EMX_MEY and Size_Range columns
    df_formatted_final['EMX_MEY'] = df_formatted['EMX_MEY']
    df_formatted_final['Size_Range'] = df_formatted['Size_Range']

    # Copy data columns to their new positions
    for col_name in data_columns:
        df_formatted_final[col_name] = df_formatted[col_name]

    # Create workbook using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample_Matrix"

    # Write data to worksheet
    for r in dataframe_to_rows(df_formatted_final, index=False, header=True):
        ws.append(r)

    # Define styles for each classification
    style_c = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Light blue
    style_a = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    style_b = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
    style_d = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")  # Grey

    # Apply conditional formatting for each classification
    # Get the data range (excluding headers)
    for row_idx in range(2, len(df_formatted_final) + 2):  # Start from row 2 (after header)
        for col_idx in range(1, len(df_formatted_final.columns) + 1):
            # Skip EMX_MEY and Size_Range columns and empty columns
            col_name = df_formatted_final.columns[col_idx - 1]
            if col_name in ["EMX_MEY", "Size_Range"] or col_name == "":
                continue
            
            # Get cell value
            cell_value = df_formatted_final.iloc[row_idx - 2, col_idx - 1]
            
            if pd.notna(cell_value):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell_value == "c":
                    cell.fill = style_c
                elif cell_value == "a":
                    cell.fill = style_a
                elif cell_value == "b":
                    cell.fill = style_b
                elif cell_value == "d":
                    cell.fill = style_d

    # Save the workbook with updated filename
    output_filename = os.path.join(output_file_path, f"Sample_Size_Matrix_Colored_{emx_mey_label}.xlsx")

    # Save workbook
    wb.save(output_filename)
    print(f"Excel file saved to: {output_filename}")
    
    return output_filename


if __name__ == "__main__":
    # Default paths for when running as main script
    input_file_path = r"C:/Users/bilya/Desktop/bibicode/R_course/GeneMapper_analysis/Input files"
    config_file_path = r"C:/Users/bilya/Desktop/bibicode/R_course/GeneMapper_analysis/Config files"
    output_file_path = r"C:/Users/bilya/Desktop/bibicode/R_course/GeneMapper_analysis/Output files"
    
    # Run the analysis
    result = gene_mapper_analysis(input_file_path, config_file_path, output_file_path, X=4, Y=8)
    print(f"Analysis complete. Output saved to: {result}")
